"""
Extract journal names from dataset_raw Excel files and merge them into
the data/ CSV files in the existing format.

CSV format:
journal_title,issn,eissn,subject_area,subject_category,publisher,oa_type,
sjr_quartile,sjr_score,impact_factor,h_index,scope
"""

import csv
import os
import re
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE, "dataset_raw")
DATA_DIR = os.path.join(BASE, "data")


def read_existing_csv(path):
    """Read existing CSV and return (headers, rows, set_of_titles)."""
    if not os.path.exists(path):
        headers = [
            "journal_title", "issn", "eissn", "subject_area",
            "subject_category", "publisher", "oa_type", "sjr_quartile",
            "sjr_score", "impact_factor", "h_index", "scope"
        ]
        return headers, [], set()

    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = list(reader)
        titles = {r.get("journal_title", "").strip().lower() for r in rows}
    return headers, rows, titles


def write_csv(path, headers, rows):
    """Write rows to CSV."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def clean(val):
    """Clean cell value."""
    if val is None:
        return ""
    s = str(val).replace("\xa0", " ").strip()
    return s


def map_oa_type(raw):
    """Map various OA type labels to standard format."""
    raw = clean(raw).lower()
    if "gold" in raw or "apc" in raw or raw == "oa":
        return "OA"
    elif "hybrid" in raw or "hibrit" in raw:
        return "Hybrid"
    elif "s2o" in raw or "subscribe" in raw:
        return "Hybrid"
    return "Hybrid"


def map_quartile(raw):
    """Map quartile values."""
    raw = clean(raw).upper()
    if not raw:
        return ""
    # Handle numeric values (1, 2, 3, 4)
    if raw in ("1", "2", "3", "4"):
        return f"Q{raw}"
    # Already formatted as Q1, Q2, etc.
    m = re.match(r"Q([1-4])", raw)
    if m:
        return f"Q{m.group(1)}"
    return raw


# ══════════════════════════════════════════════════════════════════════════════
# 1) ACS — "2026 ACS Publications Summary for Information Professionals (1).xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_acs():
    xlsx = os.path.join(RAW_DIR, "2026 ACS Publications Summary for Information Professionals (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "acs.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Journal Vitals"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
        # Columns: _, Coden, Publication, FirstYear, OA Options, Access, FirstAvail,
        #          URL, Issues, PrintISSN, PrintSvc, WebISSN, ...
        title = clean(row[2]) if len(row) > 2 else ""
        if not title or title.lower() in existing:
            continue

        issn = clean(row[9]) if len(row) > 9 else ""
        eissn = clean(row[11]) if len(row) > 11 else ""

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": "Chemistry",
            "subject_category": "Chemistry",
            "publisher": "ACS",
            "oa_type": "Hybrid",
            "sjr_quartile": "",
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"ACS: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 2) Cambridge UP — "2026-CUP-Fonlanan-dergi-listesi-AE.xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_cambridge():
    xlsx = os.path.join(RAW_DIR, "2026-CUP-Fonlanan-dergi-listesi-AE.xlsx")
    csv_path = os.path.join(DATA_DIR, "cambridge.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["AE dergi listesi"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: Title, Open Access, Print ISSN, Online ISSN
        title = clean(row[0]) if len(row) > 0 else ""
        if not title or title.lower() in existing:
            continue

        oa_raw = clean(row[1]) if len(row) > 1 else ""
        issn = clean(row[2]) if len(row) > 2 else ""
        eissn = clean(row[3]) if len(row) > 3 else ""

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": "",
            "subject_category": "",
            "publisher": "Cambridge UP",
            "oa_type": map_oa_type(oa_raw),
            "sjr_quartile": "",
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"Cambridge: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 3) Oxford UP — "OUP_fonlanan_dergi_listesi (1).xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_oxford():
    xlsx = os.path.join(RAW_DIR, "OUP_fonlanan_dergi_listesi (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "oxford.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sayfa1"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: Journal Title, Open Access Model, ONLINE_ISSN, PRINT_ISSN,
        #          Medicine, Life Sciences, Humanities, Maths & Physical Sciences,
        #          Law, Clinical Medicine, Economics & Finance, Policy,
        #          Religion & Philosophy, Social Sciences, Community College,
        #          Social Sciences & Humanities, HSS, STM
        title = clean(row[0]) if len(row) > 0 else ""
        if not title or title.lower() in existing:
            continue

        oa_raw = clean(row[1]) if len(row) > 1 else ""
        eissn = clean(row[2]) if len(row) > 2 else ""
        issn = clean(row[3]) if len(row) > 3 else ""

        # Determine subject area from columns 4-17
        subject_cols = [
            (4, "Medicine"), (5, "Life Sciences"), (6, "Humanities"),
            (7, "Mathematics & Physical Sciences"), (8, "Law"),
            (9, "Clinical Medicine"), (10, "Economics & Finance"),
            (11, "Policy"), (12, "Religion & Philosophy"),
            (13, "Social Sciences"), (14, "Community College"),
        ]
        subject_area = ""
        for col_idx, area_name in subject_cols:
            if len(row) > col_idx and row[col_idx] is not None and clean(row[col_idx]):
                subject_area = area_name
                break

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": subject_area,
            "subject_category": subject_area,
            "publisher": "Oxford UP",
            "oa_type": map_oa_type(oa_raw),
            "sjr_quartile": "",
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"Oxford: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 4) SAGE — "Sage Premier Journals Publish List 2026 (Hybrids) (1).xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_sage():
    xlsx = os.path.join(RAW_DIR, "Sage Premier Journals Publish List 2026 (Hybrids) (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "sage.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: ACCT_Ref_Code, Publication Title, ISSN, EISSN, URL,
        #          Sales Discipline, Sub Discipline, Ranked/Awarded, Imprint,
        #          Addition Type, 1st Year at SAGE, ...
        title = clean(row[1]) if len(row) > 1 else ""
        if not title or title.lower() in existing:
            continue

        issn = clean(row[2]) if len(row) > 2 else ""
        eissn = clean(row[3]) if len(row) > 3 else ""
        subject_area = clean(row[5]) if len(row) > 5 else ""
        sub_discipline = clean(row[6]) if len(row) > 6 else ""
        oa_raw = clean(row[14]) if len(row) > 14 else "Hybrid"

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": subject_area,
            "subject_category": sub_discipline,
            "publisher": "SAGE",
            "oa_type": map_oa_type(oa_raw),
            "sjr_quartile": "",
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"SAGE: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 5) Wiley — "TUBITAK_Tarafindan_Fonlanan-Wiley_AE_DergiListesi...xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_wiley():
    xlsx = os.path.join(RAW_DIR, "TUBITAK_Tarafindan_Fonlanan-Wiley_AE_DergiListesi_31_12_2026_tarihine-kadar_gecerlidir (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "wiley.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        # Row 2 is header: Sıra No, eISSN, WOS Index, Dergi Adı, Derginin Modeli,
        #                   Q Değeri 2024, IF Değeri 2024, DOI, URL,
        #                   Genel Konu, Özel Konu, Yayıncı, Not
        title = clean(row[3]) if len(row) > 3 else ""
        if not title or title.lower() in existing:
            continue

        eissn = clean(row[1]) if len(row) > 1 else ""
        oa_raw = clean(row[4]) if len(row) > 4 else ""
        q_val = row[5] if len(row) > 5 else ""
        if_val = row[6] if len(row) > 6 else ""
        subject_area = clean(row[9]) if len(row) > 9 else ""
        sub_category = clean(row[10]) if len(row) > 10 else ""

        new_row = {
            "journal_title": title,
            "issn": "",
            "eissn": eissn,
            "subject_area": subject_area,
            "subject_category": sub_category,
            "publisher": "Wiley",
            "oa_type": map_oa_type(oa_raw),
            "sjr_quartile": map_quartile(str(q_val)) if q_val else "",
            "sjr_score": "",
            "impact_factor": str(if_val) if if_val else "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"Wiley: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 6) Springer Nature — "sn_26_ae_dergi_listesi-v4 (1) (1) (1).xlsx"
# ══════════════════════════════════════════════════════════════════════════════
def extract_springer_nature():
    xlsx = os.path.join(RAW_DIR, "sn_26_ae_dergi_listesi-v4 (1) (1) (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "springer_nature.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sayfa1"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: No, Journal Title, eISSN, Quartile, Index, ISSN,
        #          Product Family, Publishing Model
        title = clean(row[1]) if len(row) > 1 else ""
        if not title or title.lower() in existing:
            continue

        eissn = clean(row[2]) if len(row) > 2 else ""
        quartile = clean(row[3]) if len(row) > 3 else ""
        issn = clean(row[5]) if len(row) > 5 else ""
        pub_model = clean(row[7]) if len(row) > 7 else ""

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": "",
            "subject_category": "",
            "publisher": "Springer Nature",
            "oa_type": map_oa_type(pub_model),
            "sjr_quartile": map_quartile(quartile),
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"Springer Nature: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 7) BSP (Bentham Science) — "BSP Active title list (1).xlsx"
#    → New publisher, create bsp.csv and add to publishers.json
# ══════════════════════════════════════════════════════════════════════════════
def extract_bsp():
    xlsx = os.path.join(RAW_DIR, "BSP Active title list (1).xlsx")
    csv_path = os.path.join(DATA_DIR, "bsp.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: Title, ISSN-Print, ISSN-Online, Journal Type, Subject field,
        #          Journal Description, CiteScore, 2024IF, URL, Publisher,
        #          First Year Published, Last Year Published, JCR, ESI,
        #          JCR Partition, SCOPUS, CAS, MEDLINE, EI Compendex, Title change
        title = clean(row[0]) if len(row) > 0 else ""
        if not title or title.lower() in existing:
            continue

        issn = clean(row[1]) if len(row) > 1 else ""
        eissn = clean(row[2]) if len(row) > 2 else ""
        oa_raw = clean(row[3]) if len(row) > 3 else ""
        subject = clean(row[4]) if len(row) > 4 else ""
        desc = clean(row[5]) if len(row) > 5 else ""
        if_val = row[7] if len(row) > 7 else ""
        quartile = clean(row[14]) if len(row) > 14 else ""

        # Clean description - remove leading title and \xa0
        desc = desc.replace("\xa0", " ").strip()

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": subject,
            "subject_category": subject,
            "publisher": "Bentham Science",
            "oa_type": map_oa_type(oa_raw),
            "sjr_quartile": map_quartile(quartile),
            "sjr_score": "",
            "impact_factor": str(if_val) if if_val else "",
            "h_index": "",
            "scope": desc,
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"BSP (Bentham Science): Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# 8) De Gruyter — "De Gruyter Complete Collection 2026 Title List (2).xlsx"
#    → New publisher, create degruyter.csv and add to publishers.json
# ══════════════════════════════════════════════════════════════════════════════
def extract_degruyter():
    xlsx = os.path.join(RAW_DIR, "De Gruyter Complete Collection 2026 Title List (2).xlsx")
    csv_path = os.path.join(DATA_DIR, "degruyter.csv")
    headers, rows, existing = read_existing_csv(csv_path)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sheet1"]

    added = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Columns: Price List, Title, Print-ISSN, Online-ISSN, 2026 Volume Number,
        #          2026 Issues per Year, Comment, URL, Subject Area, Publisher,
        #          Main Language, Publishing Model, EUR, USD
        title = clean(row[1]) if len(row) > 1 else ""
        if not title or title.lower() in existing:
            continue

        issn = clean(row[2]) if len(row) > 2 else ""
        eissn = clean(row[3]) if len(row) > 3 else ""
        subject = clean(row[8]) if len(row) > 8 else ""
        pub_model = clean(row[11]) if len(row) > 11 else ""

        new_row = {
            "journal_title": title,
            "issn": issn,
            "eissn": eissn,
            "subject_area": subject,
            "subject_category": subject,
            "publisher": "De Gruyter",
            "oa_type": map_oa_type(pub_model),
            "sjr_quartile": "",
            "sjr_score": "",
            "impact_factor": "",
            "h_index": "",
            "scope": "",
        }
        rows.append(new_row)
        existing.add(title.lower())
        added += 1

    wb.close()
    write_csv(csv_path, headers, rows)
    print(f"De Gruyter: Added {added} journals (total: {len(rows)})")


# ══════════════════════════════════════════════════════════════════════════════
# Update publishers.json — add BSP and De Gruyter
# ══════════════════════════════════════════════════════════════════════════════
def update_publishers_json():
    import json
    json_path = os.path.join(DATA_DIR, "publishers.json")

    with open(json_path, "r", encoding="utf-8") as f:
        publishers = json.load(f)

    existing_ids = {p["id"] for p in publishers}

    new_publishers = []

    if "bsp" not in existing_ids:
        new_publishers.append({
            "id": "bsp",
            "name": "Bentham Science",
            "short": "BSP",
            "csv_file": "bsp.csv",
            "color": "#8B0000",
            "bg_color": "#FFF0F0",
            "journal_count": 150,
            "description": "Bentham Science Publishers — STM journal publisher covering pharmaceutical, biomedical & engineering sciences",
            "url": "https://www.benthamscience.com",
            "oa_url": "#",
            "agreement_period": "2024–2026",
            "quota": None,
            "active": True,
            "discount": None,
            "bmc_discount": None
        })

    if "degruyter" not in existing_ids:
        new_publishers.append({
            "id": "degruyter",
            "name": "De Gruyter",
            "short": "DG",
            "csv_file": "degruyter.csv",
            "color": "#1D1D1B",
            "bg_color": "#F5F5F5",
            "journal_count": 450,
            "description": "De Gruyter — independent academic publisher covering humanities, social sciences, STEM & law",
            "url": "https://www.degruyter.com",
            "oa_url": "#",
            "agreement_period": "2024–2026",
            "quota": None,
            "active": True,
            "discount": None,
            "bmc_discount": None
        })

    if new_publishers:
        publishers.extend(new_publishers)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(publishers, f, indent=2, ensure_ascii=False)
        print(f"publishers.json: Added {len(new_publishers)} new publishers ({', '.join(p['name'] for p in new_publishers)})")
    else:
        print("publishers.json: No new publishers to add")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("Extracting journals from dataset_raw -> data/")
    print("=" * 60)

    extract_acs()
    extract_cambridge()
    extract_oxford()
    extract_sage()
    extract_wiley()
    extract_springer_nature()
    extract_bsp()
    extract_degruyter()
    update_publishers_json()

    print("\n" + "=" * 60)
    print("Done! All journal lists updated.")
    print("=" * 60)
