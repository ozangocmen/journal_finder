"""Inspect first 2 Excel files in dataset_raw to understand their structure."""
import openpyxl
import os

raw_dir = os.path.join(os.path.dirname(__file__), "dataset_raw")

for fname in sorted(os.listdir(raw_dir)):
    if fname.startswith("2026 ACS") or fname.startswith("2026-CUP") or fname.startswith("BSP"):
        fpath = os.path.join(raw_dir, fname)
        print(f"\n{'='*80}")
        print(f"FILE: {fname}")
        print(f"{'='*80}")
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\n  Sheet: '{sheet_name}'")
                for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True)):
                    print(f"    Row {i}: {row}")
                print(f"    ... (max_row={ws.max_row}, max_col={ws.max_column})")
            wb.close()
        except Exception as e:
            print(f"  ERROR: {e}")
